# backend/api_routes.py
"""
API Routes for FitTrack Backend
RESTful API endpoints with input validation and error handling
"""
import os
import io
import csv
import datetime
from flask import Blueprint, jsonify, request, url_for, redirect, current_app
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash

from backend.app import db, logger
from backend.database_models import User, Workout, WorkoutExercise
from flask import g


def _json_err(message, code=400):
    payload = {'ok': False, 'error': message}
    rid = getattr(g, 'request_id', None)
    if rid:
        payload['request_id'] = rid
    return jsonify(payload), code

# Create blueprint
api_bp = Blueprint('api', __name__)


# ============================================================================
# AUTHENTICATION & USER MANAGEMENT
# ============================================================================

@api_bp.route('/register', methods=['POST'])
def register():
    """Register a new user"""
    try:
        data = request.get_json() or {}
        
        # Validation
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return _json_err('Uživatelské jméno a heslo jsou povinné', 400)
        
        if len(username) < 3:
            return _json_err('Uživatelské jméno musí mít alespoň 3 znaky', 400)
        
        if len(password) < 8:
            return _json_err('Heslo musí mít alespoň 8 znaků', 400)
        
        # Check if user exists
        if User.query.filter_by(username=username).first():
            return _json_err('Uživatelské jméno již existuje', 400)
        
        # Create user
        new_user = User(
            username=username,
            password=generate_password_hash(password, method='pbkdf2:sha256')
        )
        db.session.add(new_user)
        db.session.commit()
        
        logger.info(f'New user registered: {username}')
        return jsonify({'ok': True, 'message': 'Registration successful'})
    
    except Exception as e:
        logger.error(f'Registration error: {str(e)}')
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Došlo k chybě při registraci. Zkuste to prosím znovu.'}), 500


@api_bp.route('/login', methods=['POST', 'GET'])
def login():
    """User login"""
    if request.method == 'GET':
        # Flask-Login redirect - return JSON response for API
        return jsonify({'ok': False, 'error': 'Authentication required', 'login_required': True}), 401
    
    try:
        data = request.get_json() or {}
        
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({'ok': False, 'error': 'Uživatelské jméno a heslo jsou povinné'}), 400
        
        # Check for admin
        admin_password = os.getenv('ADMIN_PASSWORD', 'Admin&4')
        if username.lower() == 'admin' and password == admin_password:
            admin = User.query.filter_by(username='admin').first()
            if not admin:
                admin = User(
                    username='admin',
                    password=generate_password_hash(password, method='pbkdf2:sha256')
                )
                db.session.add(admin)
                db.session.commit()
            login_user(admin)
            logger.info('Admin logged in')
            return jsonify({'ok': True, 'message': 'Logged in as admin', 'is_admin': True})
        
        # Regular user login
        user = User.query.filter_by(username=username).first()
        if not user or not check_password_hash(user.password, password):
            logger.warning(f'Failed login attempt for: {username}')
            return jsonify({'ok': False, 'error': 'Nesprávné uživatelské jméno nebo heslo'}), 401
        
        login_user(user)
        logger.info(f'User logged in: {username}')
        return jsonify({'ok': True, 'message': 'Login successful', 'is_admin': False})
    
    except Exception as e:
        logger.error(f'Login error: {str(e)}')
        return jsonify({'ok': False, 'error': 'Došlo k chybě při přihlašování. Zkuste to prosím znovu.'}), 500


# REMOVED DUPLICATE: /oauth/session route (kept the better one at line ~575)


@api_bp.route('/logout', methods=['POST'])
@login_required
def logout():
    """User logout"""
    username = current_user.username
    logout_user()
    logger.info(f'User logged out: {username}')
    return jsonify({'ok': True, 'message': 'Logged out successfully'})


@api_bp.route('/me', methods=['GET'])
@login_required
def get_current_user():
    """Get current user information"""
    try:
        user_data = current_user.to_dict(include_sensitive=True)
        return jsonify({'ok': True, 'user': user_data})
    except Exception as e:
        logger.error(f'Error fetching user data: {str(e)}')
        return jsonify({'ok': False, 'error': 'Failed to fetch user data'}), 500


@api_bp.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """Get or update user profile"""
    if request.method == 'GET':
        try:
            return jsonify({
                'ok': True,
                'profile': {
                    'age': current_user.age,
                    'height_cm': current_user.height_cm,
                    'weight_kg': current_user.weight_kg,
                    'profile_completed': current_user.profile_completed
                }
            })
        except Exception as e:
            logger.error(f'Error fetching profile: {str(e)}')
            return jsonify({'ok': False, 'error': 'Failed to fetch profile'}), 500
    
    # POST - Update profile
    try:
        data = request.get_json() or {}
        
        age = data.get('age')
        height = data.get('height_cm') or data.get('height')
        weight = data.get('weight_kg') or data.get('weight')
        
        # Validation
        if age is None or height is None or weight is None:
            return jsonify({'ok': False, 'error': 'Age, height_cm, and weight_kg are required'}), 400
        
        age = int(age)
        height = float(height)
        weight = float(weight)
        
        if not (1 <= age <= 120):
            return jsonify({'ok': False, 'error': 'Age must be between 1 and 120'}), 400
        if not (50 <= height <= 300):
            return jsonify({'ok': False, 'error': 'Height must be between 50 and 300 cm'}), 400
        if not (20 <= weight <= 500):
            return jsonify({'ok': False, 'error': 'Weight must be between 20 and 500 kg'}), 400
        
        # Update user
        user = User.query.get(current_user.id)
        user.age = age
        user.height_cm = height
        user.weight_kg = weight
        db.session.commit()
        
        logger.info(f'Profile updated for user: {current_user.username}')
        return jsonify({'ok': True, 'message': 'Profile updated successfully'})
    
    except ValueError:
        return jsonify({'ok': False, 'error': 'Invalid numeric values'}), 400
    except Exception as e:
        logger.error(f'Profile update error: {str(e)}')
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Profile update failed'}), 500


# ============================================================================
# WORKOUT MANAGEMENT
# ============================================================================

@api_bp.route('/workouts', methods=['GET'])
@login_required
def get_workouts():
    """Get all workouts for current user"""
    try:
        workouts = Workout.query.filter_by(user_id=current_user.id)\
            .order_by(Workout.date.desc()).all()
        
        workouts_data = [workout.to_dict() for workout in workouts]
        return jsonify({'ok': True, 'workouts': workouts_data})
    
    except Exception as e:
        logger.error(f'Error fetching workouts: {str(e)}')
        return jsonify({'ok': False, 'error': 'Failed to fetch workouts'}), 500


@api_bp.route('/workouts/<int:workout_id>', methods=['GET'])
@login_required
def get_workout_detail(workout_id):
    """Get detailed information about a specific workout"""
    try:
        workout = Workout.query.filter_by(id=workout_id, user_id=current_user.id).first()
        
        if not workout:
            return jsonify({'ok': False, 'error': 'Workout not found'}), 404
        
        workout_data = workout.to_dict(include_exercises=True)
        return jsonify({'ok': True, 'workout': workout_data})
    
    except Exception as e:
        logger.error(f'Error fetching workout detail: {str(e)}')
        return jsonify({'ok': False, 'error': 'Failed to fetch workout'}), 500


@api_bp.route('/workouts', methods=['POST'])
@login_required
def create_workout():
    """Create a new workout"""
    try:
        data = request.get_json() or {}
        
        # Parse date
        date_str = data.get('date')
        try:
            workout_date = datetime.date.fromisoformat(date_str) if date_str else datetime.date.today()
        except (ValueError, TypeError):
            return jsonify({'ok': False, 'error': 'Invalid date format (use YYYY-MM-DD)'}), 400
        
        note = data.get('note', '')
        exercises = data.get('exercises', [])
        
        # Create workout
        workout = Workout(
            user_id=current_user.id,
            date=workout_date,
            note=note
        )
        db.session.add(workout)
        db.session.flush()  # Get workout.id before adding exercises
        
        # Add exercises
        for ex_data in exercises:
            if not ex_data.get('name'):
                continue
            
            exercise = WorkoutExercise(
                workout_id=workout.id,
                name=ex_data['name'],
                sets=int(ex_data.get('sets', 3)),
                reps=int(ex_data.get('reps', 10)),
                weight=float(ex_data['weight']) if ex_data.get('weight') else None
            )
            db.session.add(exercise)
        
        db.session.commit()
        
        logger.info(f'Workout created: {workout.id} for user {current_user.username}')
        return jsonify({'ok': True, 'id': workout.id}), 201
    
    except ValueError as e:
        return jsonify({'ok': False, 'error': f'Invalid input: {str(e)}'}), 400
    except Exception as e:
        logger.error(f'Error creating workout: {str(e)}')
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Failed to create workout'}), 500


@api_bp.route('/workouts/<int:workout_id>', methods=['DELETE'])
@login_required
def delete_workout(workout_id):
    """Delete a workout"""
    try:
        workout = Workout.query.filter_by(id=workout_id, user_id=current_user.id).first()
        
        if not workout:
            return jsonify({'ok': False, 'error': 'Workout not found'}), 404
        
        db.session.delete(workout)
        db.session.commit()
        
        logger.info(f'Workout deleted: {workout_id} by user {current_user.username}')
        return jsonify({'ok': True, 'message': 'Workout deleted successfully'})
    
    except Exception as e:
        logger.error(f'Error deleting workout: {str(e)}')
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Failed to delete workout'}), 500


# ============================================================================
# EXERCISE MANAGEMENT
# ============================================================================

@api_bp.route('/exercises/<int:workout_id>/add', methods=['POST'])
@login_required
def add_exercise(workout_id):
    """Add an exercise to a workout"""
    try:
        workout = Workout.query.filter_by(id=workout_id, user_id=current_user.id).first()
        
        if not workout:
            return jsonify({'ok': False, 'error': 'Workout not found'}), 404
        
        data = request.get_json() or {}
        name = data.get('name', '').strip()
        
        if not name:
            return jsonify({'ok': False, 'error': 'Exercise name is required'}), 400
        
        exercise = WorkoutExercise(
            workout_id=workout.id,
            name=name,
            sets=int(data.get('sets', 3)),
            reps=int(data.get('reps', 10)),
            weight=float(data['weight']) if data.get('weight') else None
        )
        db.session.add(exercise)
        db.session.commit()
        
        logger.info(f'Exercise added to workout {workout_id}: {name}')
        return jsonify({'ok': True, 'id': exercise.id}), 201
    
    except ValueError as e:
        return jsonify({'ok': False, 'error': f'Invalid input: {str(e)}'}), 400
    except Exception as e:
        logger.error(f'Error adding exercise: {str(e)}')
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Failed to add exercise'}), 500


@api_bp.route('/exercises/<int:exercise_id>', methods=['DELETE'])
@login_required
def delete_exercise(exercise_id):
    """Delete an exercise"""
    try:
        exercise = WorkoutExercise.query\
            .join(Workout)\
            .filter(Workout.user_id == current_user.id, WorkoutExercise.id == exercise_id)\
            .first()
        
        if not exercise:
            return jsonify({'ok': False, 'error': 'Exercise not found'}), 404
        
        workout_id = exercise.workout_id
        db.session.delete(exercise)
        db.session.commit()
        
        logger.info(f'Exercise deleted: {exercise_id}')
        return jsonify({'ok': True, 'workout_id': workout_id})
    
    except Exception as e:
        logger.error(f'Error deleting exercise: {str(e)}')
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Failed to delete exercise'}), 500


# ============================================================================
# CATALOG & UTILITIES
# ============================================================================

@api_bp.route('/catalog', methods=['GET'])
@login_required
def get_exercise_catalog():
    """Get list of available exercises"""
    catalog = [
        'Bench press', 'Dřep', 'Mrtvý tah', 'Přítahy na hrazdě',
        'Tlaky na ramena', 'Biceps zdvih', 'Triceps kliky',
        'Výpady', 'Leg press', 'Veslování', 'Kettlebell swing', 'Plank'
    ]
    return jsonify({'ok': True, 'exercises': catalog})


@api_bp.route('/stats', methods=['GET'])
@login_required
def get_stats():
    """Get user statistics"""
    try:
        total_workouts = Workout.query.filter_by(user_id=current_user.id).count()
        
        recent_workouts = Workout.query.filter_by(user_id=current_user.id)\
            .order_by(Workout.date.desc()).limit(5).all()
        
        recent_exercises = sum(workout.exercises.count() for workout in recent_workouts)
        
        return jsonify({
            'ok': True,
            'stats': {
                'total_workouts': total_workouts,
                'recent_exercises': recent_exercises
            }
        })
    
    except Exception as e:
        logger.error(f'Error fetching stats: {str(e)}')
        return jsonify({'ok': False, 'error': 'Failed to fetch statistics'}), 500


@api_bp.route('/quickstart/<level>', methods=['POST'])
@login_required
def quickstart_workout(level):
    """Create a workout from a predefined template"""
    try:
        level = level.lower()
        
        presets = {
            'zacatecnik': {'label': 'Začátečník', 'sets': 3, 'reps': 10},
            'pokracily': {'label': 'Pokročilý', 'sets': 4, 'reps': 10},
            'expert': {'label': 'Expert', 'sets': 5, 'reps': 8}
        }
        
        if level not in presets:
            return jsonify({'ok': False, 'error': 'Invalid level (use: zacatecnik, pokracily, expert)'}), 400
        
        config = presets[level]
        
        workout = Workout(
            user_id=current_user.id,
            date=datetime.date.today(),
            note=f"Rychlý start – {config['label']}"
        )
        db.session.add(workout)
        db.session.flush()
        
        # Add default exercises
        default_exercises = ['Dřep', 'Bench press', 'Veslování']
        for name in default_exercises:
            exercise = WorkoutExercise(
                workout_id=workout.id,
                name=name,
                sets=config['sets'],
                reps=config['reps']
            )
            db.session.add(exercise)
        
        db.session.commit()
        
        logger.info(f'Quickstart workout created: {level} for user {current_user.username}')
        return jsonify({'ok': True, 'id': workout.id})
    
    except Exception as e:
        logger.error(f'Error creating quickstart workout: {str(e)}')
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'Failed to create workout'}), 500


# ============================================================================
# EXPORT
# ============================================================================

@api_bp.route('/export/csv', methods=['GET'])
@login_required
def export_csv():
    """Export user workouts to CSV"""
    try:
        si = io.StringIO()
        # Use semicolon delimiter for Czech Excel compatibility
        writer = csv.writer(si, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        
        # CSV headers (Czech)
        writer.writerow(['ID', 'Datum', 'Poznámka', 'Cvik', 'Série', 'Opakování', 'Váha (kg)'])
        
        workouts = Workout.query.filter_by(user_id=current_user.id)\
            .order_by(Workout.date.desc()).all()
        
        for workout in workouts:
            for exercise in workout.exercises:
                writer.writerow([
                    workout.id,
                    workout.date.strftime('%d.%m.%Y'),
                    workout.note or '',
                    exercise.name,
                    exercise.sets,
                    exercise.reps,
                    exercise.weight or ''
                ])
        
        csv_data = si.getvalue()
        logger.info(f'CSV export for user {current_user.username}: {len(workouts)} workouts')
        
        # Return CSV as proper response with correct headers
        from flask import Response
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename=fittrack_export.csv',
                'Content-Type': 'text/csv; charset=utf-8'
            }
        )
    
    except Exception as e:
        logger.error(f'CSV export error: {str(e)}')
        return jsonify({'ok': False, 'error': 'Export failed'}), 500


@api_bp.route('/export/excel', methods=['GET'])
@login_required
def export_excel():
    """Export user workouts to Excel with styling"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import tempfile
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Tréninky"
        
        # Headers (bez ID)
        headers = ['Datum', 'Trénink', 'Cvik', 'Série', 'Opakování', 'Váha (kg)']
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True, size=12, color="000000")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get workouts
        workouts = Workout.query.filter_by(user_id=current_user.id)\
            .order_by(Workout.date.desc()).all()
        
        # Add data rows grouped by workout
        row_num = 2
        workout_separator_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        workout_header_font = Font(bold=True, size=11, color="FFFFFF")
        
        for workout_idx, workout in enumerate(workouts):
            # Add workout header row (separator)
            workout_header = f"🏋️ {workout.note or 'Trénink'}"
            ws.append([workout.date.strftime('%d.%m.%Y'), workout_header, '', '', '', ''])
            
            # Style workout header
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = workout_separator_fill
                cell.font = workout_header_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Merge cells for workout header (columns B-F)
            ws.merge_cells(f'B{row_num}:F{row_num}')
            
            row_num += 1
            
            # Add exercises for this workout
            for exercise in workout.exercises:
                ws.append([
                    '',  # Empty datum (zobrazeno už v headeru)
                    '',  # Empty trénink
                    f"  • {exercise.name}",  # Odsazený cvik
                    exercise.sets,
                    exercise.reps,
                    exercise.weight if exercise.weight else ''
                ])
                
                # Style exercise row
                for col_num in range(1, 7):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Light background for exercises
                    cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                
                row_num += 1
            
            # Add empty row between workouts (except last)
            if workout_idx < len(workouts) - 1:
                ws.append(['', '', '', '', '', ''])
                row_num += 1
        
        # Auto-adjust column widths
        column_widths = {
            'A': 12,  # Datum
            'B': 35,  # Trénink/Poznámka
            'C': 30,  # Cvik
            'D': 8,   # Série
            'E': 12,  # Opakování
            'F': 12   # Váha
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Add borders (only to non-empty cells)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=row_num-1, min_col=1, max_col=6):
            for cell in row:
                if cell.value:  # Only add border if cell has content
                    cell.border = thin_border
        
        # Save to temporary file
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(tmp_file.name)
        tmp_file.close()
        
        # Read file and return
        with open(tmp_file.name, 'rb') as f:
            excel_data = f.read()
        
        import os
        os.unlink(tmp_file.name)
        
        logger.info(f'Excel export for user {current_user.username}: {len(workouts)} workouts')
        
        from flask import Response
        return Response(
            excel_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=fittrack_export.xlsx',
            }
        )
    
    except Exception as e:
        logger.error(f'Excel export error: {str(e)}')
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'ok': False, 'error': 'Export failed'}), 500


# ============================================================================
# ADMIN ENDPOINTS
# ============================================================================

@api_bp.route('/admin/users', methods=['GET'])
@login_required
def admin_get_users():
    """Get all users (admin only)"""
    if current_user.username != 'admin':
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
    
    try:
        users = User.query.order_by(User.id.asc()).all()
        
        users_data = []
        for user in users:
            user_info = user.to_dict()
            user_info['workout_count'] = Workout.query.filter_by(user_id=user.id).count()
            users_data.append(user_info)
        
        return jsonify({'ok': True, 'users': users_data})
    
    except Exception as e:
        logger.error(f'Admin users fetch error: {str(e)}')
        return jsonify({'ok': False, 'error': 'Failed to fetch users'}), 500


# ============================================================================
# OAUTH (GOOGLE)
# ============================================================================

# OAuth session endpoint
@api_bp.route('/oauth/session', methods=['POST'])
def oauth_session():
    """Set session for OAuth authenticated user"""
    try:
        data = request.get_json() or {}
        user_id = data.get('user_id')
        
        if not user_id:
            logger.warning('OAuth session: missing user_id')
            return jsonify({'ok': False, 'error': 'User ID required'}), 400
            
        user = User.query.get(user_id)
        if not user:
            logger.warning(f'OAuth session: user {user_id} not found')
            return jsonify({'ok': False, 'error': 'User not found'}), 404
            
        # Log the user in for this session
        login_user(user, remember=True)
        logger.info(f'OAuth session created for user: {user.username} (id={user.id})')
        
        return jsonify({
            'ok': True,
            'user': user.to_dict(include_sensitive=True)
        })
        
    except Exception as e:
        logger.error(f'OAuth session error: {str(e)}')
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'ok': False, 'error': str(e)}), 500


# Test endpoint
@api_bp.route('/test-oauth', methods=['GET'])
def test_oauth():
    """Test OAuth configuration"""
    try:
        from flask import current_app
        from backend.oauth import oauth, is_configured
        return jsonify({
            'oauth_configured': is_configured(),
            'oauth_object': str(oauth),
            'google_client_id': current_app.config.get('GOOGLE_CLIENT_ID', 'NOT SET'),
            'google_client_secret_set': bool(current_app.config.get('GOOGLE_CLIENT_SECRET'))
        })
    except Exception as e:
        return jsonify({'error': str(e)})

@api_bp.route('/google/login', methods=['GET'])
def google_login():
    """Initiate Google OAuth login - Simple version"""
    try:
        from backend.oauth import oauth, is_configured
        
        if not is_configured() or oauth is None:
            return jsonify({'ok': False, 'error': 'Google OAuth is not configured'}), 501
        
        # Simple redirect without extra complexity
        redirect_uri = url_for('api.google_callback', _external=True)
        
        # Create authorization URL manually for better control
        from urllib.parse import urlencode
        import secrets
        
        google_auth_url = 'https://accounts.google.com/o/oauth2/auth'
        state = secrets.token_urlsafe(32)
        
        params = {
            'client_id': oauth.google.client_id,
            'redirect_uri': redirect_uri,
            'scope': 'openid email profile',
            'response_type': 'code',
            'state': state,
        }
        
        auth_url = f"{google_auth_url}?{urlencode(params)}"
        
        # Store state in session for security
        from flask import session
        session['oauth_state'] = state
        
        # Return redirect response
        from flask import redirect
        return redirect(auth_url)
    
    except Exception as e:
        import traceback
        logger.error(f'Google login error: {str(e)}')
        logger.error(f'Traceback: {traceback.format_exc()}')
        return jsonify({'ok': False, 'error': f'OAuth failed: {str(e)}'}), 500


@api_bp.route('/google/callback', methods=['GET'])
def google_callback():
    """Handle Google OAuth callback - Manual implementation"""
    # Log immediately to see if we even get here
    print('=== GOOGLE CALLBACK STARTED ===', flush=True)
    
    try:
        from flask import request, session
        
        # Get the authorization code and state from query params
        code = request.args.get('code')
        state = request.args.get('state')
        error = request.args.get('error')
        
        print(f'Google callback params: code={bool(code)}, state={bool(state)}, error={error}', flush=True)
        logger.info(f'Google callback received: code={bool(code)}, state={bool(state)}, error={error}')
        
        # Check for OAuth errors
        if error:
            print(f'=== ERROR IN PARAMS: {error} ===', flush=True)
            frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
            return redirect(f'{frontend_url}/?auth=error&msg={error}')
        
        if not code:
            print(f'=== NO CODE ===', flush=True)
            frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
            return redirect(f'{frontend_url}/?auth=error&msg=No+authorization+code+received')
        
        # Verify state parameter (CSRF protection)
        print(f'=== CHECKING STATE ===', flush=True)
        
        try:
            stored_state = session.get('oauth_state')
            print(f'=== STORED STATE: {stored_state}, RECEIVED STATE: {state} ===', flush=True)
        except Exception as e:
            print(f'=== EXCEPTION GETTING SESSION: {e} ===', flush=True)
            raise
        
        if not stored_state or stored_state != state:
            print(f'=== STATE MISMATCH! ===', flush=True)
            frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
            return redirect(f'{frontend_url}/?auth=error&msg=Invalid+state+parameter')
        
        print(f'=== STATE OK, CLEARING ===', flush=True)
        # Clear the state from session
        session.pop('oauth_state', None)
        
        # Exchange authorization code for access token
        import requests
        from backend.oauth import oauth
        
        print(f'=== OAUTH OBJECT: {oauth} ===', flush=True)
        print(f'=== HAS GOOGLE: {hasattr(oauth, "google") if oauth else False} ===', flush=True)
        
        logger.info('Exchanging authorization code for token...')
        
        # Check if OAuth is configured
        if oauth is None:
            logger.error('OAuth not configured!')
            frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
            return redirect(f'{frontend_url}/?auth=error&msg=OAuth+not+configured')
        
        if not hasattr(oauth, 'google'):
            logger.error('OAuth google client not registered!')
            frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
            return redirect(f'{frontend_url}/?auth=error&msg=OAuth+google+not+configured')
        
        token_url = 'https://oauth2.googleapis.com/token'
        redirect_uri = url_for('api.google_callback', _external=True)
        
        logger.info(f'=== REDIRECT_URI: {redirect_uri} ===')
        
        token_data = {
            'client_id': oauth.google.client_id,
            'client_secret': oauth.google.client_secret,
            'code': code,
            'grant_type': 'authorization_code',
            'redirect_uri': redirect_uri,
        }
        
        logger.info(f'=== TOKEN REQUEST: client_id={token_data["client_id"]}, redirect_uri={token_data["redirect_uri"]}, grant_type={token_data["grant_type"]}, code_length={len(code)} ===')
        
        token_response = requests.post(token_url, data=token_data)
        token_json = token_response.json()
        
        logger.info(f'Token response status: {token_response.status_code}')
        logger.info(f'=== TOKEN RESPONSE BODY: {token_json} ===')
        
        if not token_response.ok:
            error_msg = token_json.get('error_description', 'Token exchange failed')
            logger.error(f'Token exchange failed: {error_msg}')
            frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
            return redirect(f'{frontend_url}/?auth=error&msg={error_msg}')
        
        access_token = token_json.get('access_token')
        id_token = token_json.get('id_token')
        
        if not access_token:
            frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
            return redirect(f'{frontend_url}/?auth=error&msg=No+access+token+received')
        
        # Get user info from Google
        userinfo_url = f'https://www.googleapis.com/oauth2/v2/userinfo?access_token={access_token}'
        userinfo_response = requests.get(userinfo_url)
        userinfo = userinfo_response.json()
        
        print(f'=== USERINFO RESPONSE: status={userinfo_response.status_code} ===', flush=True)
        
        if not userinfo_response.ok:
            logger.error(f'Failed to get userinfo: {userinfo}')
            frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
            return redirect(f'{frontend_url}/?auth=error&msg=Failed+to+get+user+info')
        
        sub = userinfo.get('id')  # Google uses 'id' field
        email = userinfo.get('email')
        name = userinfo.get('name') or (email.split('@')[0] if email else f'user_{sub[:6]}')
        
        logger.info(f'Google user info: email={email}, name={name}, sub={sub[:10] if sub else None}...')
        
        # Find or create user
        user = User.query.filter(
            (User.oauth_provider == 'google') & (User.oauth_sub == sub)
        ).first()
        
        if not user and email:
            user = User.query.filter_by(email=email).first()
        
        if not user:
            logger.info(f'Creating new user for Google OAuth: {name}')
            user = User(
                username=name,
                email=email,
                oauth_provider='google',
                oauth_sub=sub,
                password=generate_password_hash(os.urandom(16).hex())
            )
            db.session.add(user)
            db.session.commit()
            logger.info(f'New user created: {user.id}')
        
        login_user(user)
        logger.info(f'Google OAuth login successful: {user.username} (id={user.id})')
        print(f'=== LOGIN_USER CALLED for user {user.id} ===', flush=True)
        
        frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
        redirect_url = f'{frontend_url}/?auth=success&user_id={user.id}'
        print(f'=== REDIRECTING TO: {redirect_url} ===', flush=True)
        logger.info(f'Redirecting to: {redirect_url}')
        return redirect(redirect_url)
    
    except Exception as e:
        print(f'=== GOOGLE CALLBACK EXCEPTION: {str(e)} ===', flush=True)
        logger.error(f'Google callback error: {str(e)}')
        import traceback
        traceback_str = traceback.format_exc()
        print(f'Traceback:\n{traceback_str}', flush=True)
        logger.error(f'Traceback: {traceback_str}')
        frontend_url = current_app.config.get('FRONTEND_URL', 'http://localhost:8501')
        return redirect(f'{frontend_url}/?auth=error&msg={str(e)}')


@api_bp.route('/check_username', methods=['GET'])
def check_username():
    """Check if a username is available (quick client-side hint)."""
    try:
        uname = request.args.get('username', '').strip()
        if not uname:
            return _json_err('username parameter is required', 400)

        exists = User.query.filter_by(username=uname).first() is not None
        payload = {'ok': True, 'available': not exists}
        # attach request id if present
        rid = getattr(g, 'request_id', None)
        if rid:
            payload['request_id'] = rid
        return jsonify(payload)
    except Exception as e:
        logger.error(f'Username check error: {str(e)}')
        return _json_err('Could not check username', 500)
